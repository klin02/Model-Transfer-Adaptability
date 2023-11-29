from torch.serialization import load
from model import *
from extract_ratio import *
from utils import *

import gol
import openpyxl
import sys
import argparse
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torchvision import datasets, transforms
from torchvision.transforms.functional import InterpolationMode
import torch.utils.bottleneck as bn
import os
import os.path as osp
from torch.utils.tensorboard import SummaryWriter


def direct_quantize(model, test_loader,device):
    for i, (data, target) in enumerate(test_loader, 1):
        data = data.to(device)
        output = model.quantize_forward(data).cpu()
        if i % 500 == 0:
            break
    print('direct quantization finish')


def full_inference(model, test_loader, device):
    correct = 0
    for i, (data, target) in enumerate(test_loader, 1):
        data = data.to(device)
        output = model(data).cpu()
        pred = output.argmax(dim=1, keepdim=True)
        # print(pred)
        correct += pred.eq(target.view_as(pred)).sum().item()
    print('\nTest set: Full Model Accuracy: {:.2f}%'.format(100. * correct / len(test_loader.dataset)))
    return 100. * correct / len(test_loader.dataset)


def quantize_inference(model, test_loader, device):
    correct = 0
    for i, (data, target) in enumerate(test_loader, 1):
        data = data.to(device)
        output = model.quantize_inference(data).cpu()
        pred = output.argmax(dim=1, keepdim=True)
        correct += pred.eq(target.view_as(pred)).sum().item()
    print('Test set: Quant Model Accuracy: {:.2f}%'.format(100. * correct / len(test_loader.dataset)))
    return 100. * correct / len(test_loader.dataset)


if __name__ == "__main__":
    batch_size = 32

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(device)
    train_loader = torch.utils.data.DataLoader(
        datasets.CIFAR10('../data', train=True, download=True,
                         transform=transforms.Compose([
                             transforms.Resize((32, 32), interpolation=InterpolationMode.BICUBIC),
                             transforms.RandomHorizontalFlip(),
                             transforms.ToTensor(),
                             transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))
                         ])),
        batch_size=batch_size, shuffle=True, num_workers=1, pin_memory=True
    )

    test_loader = torch.utils.data.DataLoader(
        datasets.CIFAR10('../data', train=False, transform=transforms.Compose([
            transforms.Resize((32, 32), interpolation=InterpolationMode.BICUBIC),
            transforms.ToTensor(),
            transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))
        ])),
        batch_size=batch_size, shuffle=True, num_workers=1, pin_memory=True
    )

    model = VGG_19()
    writer = SummaryWriter(log_dir='./log')
    full_file = 'ckpt/cifar10_VGG_19.pt'
    model.load_state_dict(torch.load(full_file))
    model.to(device)

    load_ptq = True
    store_ptq = False
    ptq_file_prefix = 'ckpt/cifar10_VGG_19_ptq_'

    model.eval()
    full_acc = full_inference(model, test_loader, device)
    model_fold = fold_model(model)
    
    full_params = []

    layer, par_ratio, flop_ratio = extract_ratio()
    par_ratio, flop_ratio = fold_ratio(layer, par_ratio, flop_ratio)

    for name, param in model_fold.named_parameters():
        if 'bn' in name:
            continue
        param_norm = F.normalize(param.data.cpu(),p=2,dim=-1)
        full_params.append(param_norm)
        writer.add_histogram(tag='Full_' + name + '_data', values=param.data)
    
    
    gol._init()
    quant_type_list = ['INT','POT','FLOAT']
    # quant_type_list = ['FLOAT']
    title_list = []
    js_flops_list = []
    js_param_list = []
    ptq_acc_list = []
    acc_loss_list = []

    for quant_type in quant_type_list:
        num_bit_list = numbit_list(quant_type)
        # 对一个量化类别，只需设置一次bias量化表
        # int由于位宽大，使用量化表开销过大，直接_round即可
        if quant_type != 'INT':
            bias_list = build_bias_list(quant_type)
            gol.set_value(bias_list, is_bias=True)

        for num_bits in num_bit_list:
            e_bit_list = ebit_list(quant_type,num_bits)
            for e_bits in e_bit_list:
                model_ptq = VGG_19()
                if quant_type == 'FLOAT':
                    title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                else:
                    title = '%s_%d' % (quant_type, num_bits)
                print('\nPTQ: '+title)
                title_list.append(title)

                # 设置量化表
                if quant_type != 'INT':
                    plist = build_list(quant_type, num_bits, e_bits)
                    gol.set_value(plist)

                # 判断是否需要载入
                if load_ptq is True and osp.exists(ptq_file_prefix + title + '.pt'):
                    model_ptq.quantize(quant_type,num_bits,e_bits)
                    model_ptq.load_state_dict(torch.load(ptq_file_prefix + title + '.pt'))
                    model_ptq.to(device)
                    print('Successfully load ptq model: ' + title)
                else:
                    model_ptq.load_state_dict(torch.load(full_file))
                    model_ptq.to(device)
                    model_ptq.quantize(quant_type,num_bits,e_bits)
                    model_ptq.eval()
                    direct_quantize(model_ptq, train_loader, device)
                    if store_ptq:
                        torch.save(model_ptq.state_dict(), ptq_file_prefix + title + '.pt')

                model_ptq.freeze()
                ptq_acc = quantize_inference(model_ptq, test_loader, device)
                ptq_acc_list.append(ptq_acc)
                acc_loss = (full_acc - ptq_acc) / full_acc
                acc_loss_list.append(acc_loss)
                idx = -1

                # 获取计算量/参数量下的js-div
                js_flops = 0.
                js_param = 0.
                for name, param in model_ptq.named_parameters():
                    if '.' not in name or 'bn' in name:
                        continue
                    idx = idx + 1
                    prefix = name.split('.')[0]
                    if prefix in layer:
                        layer_idx = layer.index(prefix)
                        ptq_param = param.data.cpu()
                        # 取L2范数
                        ptq_norm = F.normalize(ptq_param,p=2,dim=-1)
                        writer.add_histogram(tag=title +':'+ name + '_data', values=ptq_param)
                        js = js_div(ptq_norm,full_params[idx])
                        js = js.item()
                        if js < 0.:
                            js = 0.
                        js_flops = js_flops + js * flop_ratio[layer_idx]
                        js_param = js_param + js * par_ratio[layer_idx]
                js_flops_list.append(js_flops)
                js_param_list.append(js_param)

                print(title + ': js_flops: %f js_param: %f acc_loss: %f' % (js_flops, js_param, acc_loss))
    
    # 写入xlsx
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1,column=1,value='FP32-acc')
    worksheet.cell(row=1,column=2,value=full_acc)
    worksheet.cell(row=3,column=1,value='title')
    worksheet.cell(row=3,column=2,value='js_flops')
    worksheet.cell(row=3,column=3,value='js_param')
    worksheet.cell(row=3,column=4,value='ptq_acc')
    worksheet.cell(row=3,column=5,value='acc_loss')
    for i in range(len(title_list)):
        worksheet.cell(row=i+4, column=1, value=title_list[i])
        worksheet.cell(row=i+4, column=2, value=js_flops_list[i])
        worksheet.cell(row=i+4, column=3, value=js_param_list[i])
        worksheet.cell(row=i+4, column=4, value=ptq_acc_list[i])
        worksheet.cell(row=i+4, column=5, value=acc_loss_list[i])
    
    workbook.save('ptq_result.xlsx')
    writer.close()

    ft = open('ptq_result.txt','w')

    print('title_list:',file=ft)
    print(" ".join(title_list),file=ft)
    print('js_flops_list:',file=ft)
    print(" ".join(str(i) for i in js_flops_list), file=ft)
    print('js_param_list:',file=ft)
    print(" ".join(str(i) for i in js_param_list), file=ft)
    print('ptq_acc_list:',file=ft)
    print(" ".join(str(i) for i in ptq_acc_list), file=ft)
    print('acc_loss_list:',file=ft)
    print(" ".join(str(i) for i in acc_loss_list), file=ft)

    ft.close()

