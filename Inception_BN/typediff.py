from torch.serialization import load
from model import *
from extract_ratio import *
from utils import *

import gol
import openpyxl
import sys
import argparse
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torchvision import datasets, transforms
from torchvision.transforms.functional import InterpolationMode
import torch.utils.bottleneck as bn
import os
import os.path as osp
from torch.utils.tensorboard import SummaryWriter


def build_int_list(num_bits):
    plist = [0.]
    for i in range(0,2**(num_bits-1)): 
        # i最高到0，即pot量化最大值为1
        plist.append(i)
        plist.append(i)
    plist = torch.Tensor(list(set(plist)))
    # plist = plist.mul(1.0 / torch.max(plist))
    return plist


def js_div_diff(p_output, q_output, get_softmax=True):
    """
    Function that measures JS divergence between target and output logits:
    """
    KLDivLoss = nn.KLDivLoss(reduction='sum')
    if get_softmax:
        p_output = F.softmax(p_output)
        q_output = F.softmax(q_output)
    p_output = p_output.view(-1)
    q_output = q_output.view(-1)
    # log_mean_output = ((p_output + q_output)/2).log()
    log_mean_output = 0.
    log_mean_output += p_output.log()
    log_mean_output += (q_output / q_output.size(0)).log()
    log_mean_output /= 2.0
    return (KLDivLoss(log_mean_output, p_output) + KLDivLoss(log_mean_output, q_output))/2


if __name__ == "__main__":
    batch_size = 32

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(device)


    model = Inception_BN()
    full_file = 'ckpt/cifar10_Inception_BN.pt'
    model.load_state_dict(torch.load(full_file))
    model.to(device)

    model.eval()
    # 传入后可变
    fold_model(model)
    
    layer, par_ratio, flop_ratio = extract_ratio()
    par_ratio, flop_ratio = fold_ratio(layer, par_ratio, flop_ratio)

    full_names = []
    full_params = []

    for name, param in model.named_parameters():
        if 'conv' in name or 'fc' in name:
            full_names.append(name)
            param_norm = F.normalize(param.data.cpu(),p=2,dim=-1)
            full_params.append(param_norm)
    
    #统计每个参数对应层的参数个数
    full_par_num=[]
    for name in full_names:
        prefix = name.rsplit('.',1)[0]
        cnt = 0
        for str in full_names:
            sprefix = str.rsplit('.',1)[0]
            if prefix == sprefix:
                cnt += 1
        full_par_num.append(cnt)
    
    gol._init()
    # quant_type_list = ['INT','POT','FLOAT']
    # quant_type_list = ['INT']
    # quant_type_list = ['POT']
    quant_type_list = ['INT','POT']
    title_list = []
    js_flops_list = []
    js_param_list = []

    for quant_type in quant_type_list:
        num_bit_list = numbit_list(quant_type)
        # 对一个量化类别，只需设置一次bias量化表
        # int由于位宽大，使用量化表开销过大，直接_round即可
        if quant_type != 'INT':
            bias_list = build_bias_list(quant_type)
            gol.set_value(bias_list, is_bias=True)
        else:
            bias_list = build_int_list(16)
            gol.set_value(bias_list, is_bias=True)

        for num_bits in num_bit_list:
            e_bit_list = ebit_list(quant_type,num_bits)
            for e_bits in e_bit_list:
                if quant_type == 'FLOAT':
                    title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                else:
                    title = '%s_%d' % (quant_type, num_bits)
                print('\nDIFF: '+title)
                title_list.append(title)

                # 设置量化表
                if quant_type != 'INT':
                    plist = build_list(quant_type, num_bits, e_bits)
                    gol.set_value(plist)
                else:
                    plist = build_int_list(num_bits)
                    gol.set_value(plist)


                # 获取计算量/参数量下的js-div
                js_flops = 0.
                js_param = 0.
                for name, _ in model.named_parameters():
                    if 'conv' not in name and 'fc' not in name:
                        continue
                    if 'weight' in name:
                        plist = gol.get_value(False)
                    else:
                        plist = gol.get_value(True)    
                    prefix = name.rsplit('.',1)[0]
                    layer_idx = layer.index(prefix)
                    name_idx = full_names.index(name)

                    layer_idx = layer.index(prefix)
                    # ptq_param = param.data.cpu()
                    # 取L2范数
                    plist_norm = F.normalize(plist,p=2,dim=-1)
                    
                    print(name)
                    print(plist_norm)
                    print(full_params[name_idx])
                    js = js_div_diff(plist_norm,full_params[name_idx])
                    js /= full_par_num[name_idx]
                    js = js.item()
                    if js < 0.:
                        js = 0.
                    js_flops = js_flops + js * flop_ratio[layer_idx]
                    js_param = js_param + js * par_ratio[layer_idx]
                js_flops_list.append(js_flops)
                js_param_list.append(js_param)

                print(title + ': js_flops: %f js_param: %f ' % (js_flops, js_param))
    
    # 写入xlsx
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(row=3,column=1,value='title')
    worksheet.cell(row=3,column=2,value='js_flops')
    worksheet.cell(row=3,column=3,value='js_param')
    for i in range(len(title_list)):
        worksheet.cell(row=i+4, column=1, value=title_list[i])
        worksheet.cell(row=i+4, column=2, value=js_flops_list[i])
        worksheet.cell(row=i+4, column=3, value=js_param_list[i])
    

    ft = open('diff_result.txt','w')

    print('title_list:',file=ft)
    print(" ".join(title_list),file=ft)
    print('js_flops_list:',file=ft)
    print(" ".join(str(i) for i in js_flops_list), file=ft)
    print('js_param_list:',file=ft)
    print(" ".join(str(i) for i in js_param_list), file=ft)

    ft.close()

