from torch.serialization import load
from model import *
from extract_ratio import *
from utils import *
from lstm_utils import *

import gol
import openpyxl
import sys
import argparse
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torchvision import datasets, transforms
from torchvision.transforms.functional import InterpolationMode
import os
import os.path as osp
from torch.utils.tensorboard import SummaryWriter

import time
import math

data_path = '../data/ptb'
embed_size = 700
hidden_size = 700
lr = 22
clip = 0.25
eval_batch_size = 10
bptt = 35 #所取的串长度
dropout = 0.65
tied = True
seed = 1111
seed_gpu = 1111
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')


def direct_quantize(model, val_data, eval_batch_size, bptt):
    # hidden = model.init_hidden(eval_batch_size)
    hidden=None
    with torch.no_grad():
        for i in range(0, val_data.size(0) - 1, bptt):
            data, targets = get_batch(val_data, i, bptt)
            output, hidden = model.quantize_forward(data,hidden)
            hidden = repackage_hidden(hidden)
    print('direct quantization finish')

def full_inference(model, test_data, ntokens, eval_batch_size, bptt):
    total_loss = 0.
    lossLayer = nn.CrossEntropyLoss()
    # hidden = model.init_hidden(eval_batch_size)
    hidden=None
    with torch.no_grad():
        for i in range(0, test_data.size(0) - 1, bptt):
            data, targets = get_batch(test_data, i, bptt)
            output, hidden = model(data,hidden)
            output_flat = output.view(-1, ntokens)
            total_loss += len(data) * lossLayer(output_flat, targets).item()
            hidden = repackage_hidden(hidden)
    test_loss = total_loss / (len(test_data) - 1)
    ppl = math.exp(test_loss)
    print('\nTest set: Full Model Perplexity: {:.4f} Loss {:f}'.format(ppl,test_loss))
    return ppl

def quantize_inference(model, test_data, ntokens, eval_batch_size, bptt):
    total_loss = 0.
    lossLayer = nn.CrossEntropyLoss()
    # hidden = model.init_hidden(eval_batch_size)
    hidden=None
    # print(model.qembed1.qw)
    # print(model.qembed1.qo)
    # print(model.qlstm3.qix)
    # print(model.qlstm3.qih)
    # print(model.qlstm3.qic)
    # print(model.qlstm3.qox)
    # print(model.qlstm3.qoh)
    # print(model.qlstm3.qoc)
    with torch.no_grad():
        for i in range(0, test_data.size(0) - 1, bptt):
            data, targets = get_batch(test_data, i, bptt)
            output, hidden = model.quantize_inference(data,hidden)
            output_flat = output.view(-1, ntokens)
            total_loss += len(data) * lossLayer(output_flat, targets).item()
            hidden = repackage_hidden(hidden)
    test_loss = total_loss / (len(test_data) - 1)
    ppl = math.exp(test_loss)
    print('Test set: Quant Model Perplexity: {:.4f} Loss {:f}\n'.format(ppl,test_loss))
    return ppl


if __name__ == "__main__":
    sys.stdout = open(sys.stdout.fileno(), mode='w', buffering=1)
    
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed_gpu)

    corpus = Corpus(data_path)
    ntokens = len(corpus.dictionary)

    val_data = batchify(corpus.valid, eval_batch_size, device)
    test_data = batchify(corpus.test, eval_batch_size, device)


    load_ptq = False
    store_ptq = False

    gol._init()

    excel_path = 'ptq_result.xlsx'
    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    txt_path = 'ptq_result.txt'
    ft = open(txt_path,'w')
    
    model = Model(ntokens, embed_size, hidden_size, dropout, tied).to(device)
    full_file = 'ckpt/ptb_PTB_LSTM.pt'
    model.load_state_dict(torch.load(full_file))

    ptq_file_prefix = 'ckpt/ptb_PTB_LSTM_ptq_'

    model.eval()
    lstm_flatten(model)
    full_ppl = full_inference(model,test_data,ntokens,eval_batch_size,bptt)
    
    quant_type_list = ['INT','POT','FLOAT']
    # quant_type_list = ['FLOAT']

    title_list = []
    js_flops_list = []
    js_param_list = []
    ptq_ppl_list = []
    ppl_ratio_list = []

    for quant_type in quant_type_list:
        num_bit_list = numbit_list(quant_type)
        if quant_type != 'INT':
            bias_list = build_bias_list(quant_type)
            gol.set_value(bias_list, is_bias=True)

        for num_bits in num_bit_list:
            e_bit_list = ebit_list(quant_type,num_bits)
            for e_bits in e_bit_list:
                model_ptq = Model(ntokens, embed_size, hidden_size, dropout, tied).to(device)
                if quant_type == 'FLOAT':
                    title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                else:
                    title = '%s_%d' % (quant_type, num_bits)
                print('\nPTB_LSTM: PTQ: '+title)
                title_list.append(title)

                if quant_type != 'INT':
                    plist = build_list(quant_type, num_bits, e_bits)
                    gol.set_value(plist)
                
                model_ptq.load_state_dict(torch.load(full_file))
                model_ptq.quantize(quant_type,num_bits,e_bits)
                model_ptq.eval()
                direct_quantize(model_ptq, val_data, eval_batch_size, bptt)

                #这里的quantize_inference都用伪量化，相比forward只少了update，不需freeze
                ptq_ppl = quantize_inference(model_ptq, test_data, ntokens, eval_batch_size, bptt)
                ppl_ratio = ptq_ppl / full_ppl
                print(title+': ppl_ratio: %f'%ppl_ratio)
                ptq_ppl_list.append(ptq_ppl)
                ppl_ratio_list.append(ppl_ratio)

    worksheet = workbook.create_sheet('PTB_LSTM')
    worksheet.cell(row=1,column=1,value='FP32-ppl')
    worksheet.cell(row=1,column=2,value=full_ppl)
    worksheet.cell(row=3,column=1,value='title')
    # worksheet.cell(row=3,column=2,value='js_flops')
    # worksheet.cell(row=3,column=3,value='js_param')
    worksheet.cell(row=3,column=2,value='ptq_ppl')
    worksheet.cell(row=3,column=3,value='ppl_ratio')

    for i in range(len(title_list)):
        worksheet.cell(row=i+4,column=1,value=title_list[i])
        worksheet.cell(row=i+4,column=2,value=ptq_ppl_list[i])
        worksheet.cell(row=i+4,column=3,value=ppl_ratio_list[i])

    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    workbook.save(excel_path)

    print('PTB_LSTM',file=ft)
    print('Full_ppl: %f'%full_ppl,file=ft)
    print('title_list:',file=ft)
    print(title_list,file=ft)
    # print('js_flops_list:',file=ft)
    # print(js_flops_list, file=ft)
    # print('js_param_list:',file=ft)
    # print(js_param_list, file=ft)
    print('ptq_ppl_list:',file=ft)
    print(ptq_ppl_list, file=ft)
    print('ppl_ratio_list:',file=ft)
    print(ppl_ratio_list, file=ft)
    print("\n",file=ft)
    
    ft.close()
