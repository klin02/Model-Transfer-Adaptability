from model import Model
from dataloader import DataLoader
from utils import numbit_list, ebit_list, build_bias_list, build_list
# from generator import Generator,Generator_imagenet
import module
import gol
from robust_utils import build_gen_loader, test_autoattack, test_robust
from gen_one import GenOption

import torch.nn.functional as F
import numpy as np
import argparse
import time
import torch
import torch.nn as nn
import sys
import os
import os.path as osp
import openpyxl

    
class RobustOption(object):
    def __init__(self, args):
        self.model = args.model
        self.dataset = args.dataset
        self.batchSize = 128
        self.channels = 3
        if self.dataset == "cifar10":
            self.nClasses = 10
            self.img_size = 32
        elif self.dataset == "cifar100":
            self.nClasses = 100
            self.img_size = 32
        else:
            assert False, "invalid dataset"

        # 对抗攻击相关信息
        self.gen_iters = 100 #伪数据集生成轮数
        self.eps = 8/255
        self.steps = 10
        self.coeff = 0.1
            
        # ----------Eval options ---------------------------------------------
        #每个模型需要评估多少轮，和batchsize无关
        self.iters = 200
        # 根据加载的生成器进行调整
        self.latent_dim = 64

        # ----------Noise Options--------------------------------------------
        # 关于对边界样本扰动的噪声指标
        # self.noise_mean = 0.0
        # self.noise_sigma = 0.01

        self.adjust = args.adjust
        if self.adjust:
            self.gen_file = 'ckpt_gen/'+self.dataset+'/'+self.model+'_adjust.pt'
        else:
            self.gen_file = 'ckpt_gen/'+self.dataset+'/'+self.model+'.pt'

    # 逐个加载全精度和量化模型
    def set(self,quant,quant_type=None,num_bits=None,e_bits=None):
        self.quant = quant
        if self.quant:
            self.quant_type = quant_type
            self.num_bits = num_bits
            self.e_bits = e_bits
            if quant_type == 'FLOAT':
                title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
            else:
                title = '%s_%d' % (quant_type, num_bits)
            self.model_file = 'ckpt_quant/'+self.dataset+'/'+self.model+'/'+title+'.pt'
        else:
            self.model_file = 'ckpt_full/'+self.dataset+'/'+self.model+'.pt'

        if not osp.exists(self.model_file):
            assert False, "Empty model file"

class RobustEvaler(object):
    def __init__(self, option):
        self.settings = option
        self.set_test_loader()
        self.set_gen_loader()    # 设置全精度模型生成器的伪数据集

    def set(self,quant,quant_type=None,num_bits=None,e_bits=None):
        # 调整option非公共部分及相关结构
        self.settings.set(quant,quant_type,num_bits,e_bits)
        self.set_model()
        
    def set_test_loader(self):
        dataloader = DataLoader(self.settings.dataset,self.settings.batchSize)
        _,_,self.test_loader = dataloader.getloader()

    def set_model(self):
        self.model = Model(self.settings.model,self.settings.dataset).cuda()
        if self.settings.quant:
            self.model.quantize(self.settings.quant_type,self.settings.num_bits,self.settings.e_bits)
        self.model.load_state_dict(torch.load(self.settings.model_file))
        self.model.eval()

    # 这里直接加载gen_one训练的完整模型，避免结构不同。注意：latent_dim需要根据加载的文件进行修改
    def set_gen_loader(self):
        self.generator = torch.load(self.settings.gen_file)
        self.settings.latent_dim = self.generator.settings.latent_dim
        
        self.gen_loader = build_gen_loader(generator=self.generator,
                                           batchSize=self.settings.batchSize,
                                           iters=self.settings.gen_iters,
                                           latent_dim=self.settings.latent_dim,
                                           nClasses=self.settings.nClasses)

    def run(self):
        start_time = time.time()
        # org_acc, rob_acc = self.eval_robust()
        print('Evaluate Orgin')
        org_acc = test_robust(self.model, attack_type=None, c=self.settings.eps, num_classes=self.settings.nClasses,
                            testloader=self.test_loader)
        org_gen_acc = test_robust(self.model, attack_type=None, c=self.settings.eps, num_classes=self.settings.nClasses,
                            testloader=self.gen_loader)
        print('Evaluate FGSM:')
        fgsm_acc = test_robust(self.model, attack_type='fgsm', c=self.settings.eps, num_classes=self.settings.nClasses,
                            testloader=self.test_loader)
        fgsm_gen_acc = test_robust(self.model, attack_type='fgsm', c=self.settings.eps, num_classes=self.settings.nClasses,
                            testloader=self.gen_loader)
        print('Evaluate PGD:')
        pgd_acc = test_robust(self.model, attack_type='pgd', c=self.settings.eps, num_classes=self.settings.nClasses,
                            testloader=self.test_loader)
        pgd_gen_acc = test_robust(self.model, attack_type='pgd', c=self.settings.eps, num_classes=self.settings.nClasses,
                            testloader=self.gen_loader)   

        # 耗时太长，舍弃
        # print('Evaluate CW:')
        # test_robust(self.model, attack_type='cw', c=self.settings.coeff, num_classes=self.settings.nClasses,
        # 			testloader=self.test_loader, loss_fn=nn.CrossEntropyLoss(), req_count=10000)
        # 不兼容python3.6，且量化模型存在零梯度点
        # print('Evaluate AA:')
        # aa_acc = test_autoattack(self.model, self.test_loader, norm='Linf', eps=self.settings.eps,
        # 				version='standard', verbose=False)
        # aa_gen_acc = test_autoattack(self.model, self.gen_loader, norm='Linf', eps=self.settings.eps,
        # 				version='standard', verbose=False)

        time_interval = time.time()-start_time
        start_time = time.time()

        print('=== Time:%.2fs'%time_interval)
        return org_acc, org_gen_acc, fgsm_acc, fgsm_gen_acc, pgd_acc, pgd_gen_acc 


# 比较全精度模型和量化模型对边界扰动的耐受程度
# 度量指标：以边界样本分类为基准，考虑加上噪声后结果的改变
def main():
    #及时打印信息
    sys.stdout = open(sys.stdout.fileno(), mode='w', buffering=1)
    torch.backends.cudnn.enabled = True
    torch.backends.cudnn.benchmark = True
    parser = argparse.ArgumentParser(description='Robust Arg')
    parser.add_argument('--model', type=str)
    parser.add_argument('--dataset',type=str)

    parser.add_argument('--multi_label_num', type=int, default=2)
    # 是否使用调整后的生成器
    parser.add_argument('--adjust', action='store_true')

    args = parser.parse_args()
    print(args)

    option = RobustOption(args)

    robust_dir = 'robust_result/'+option.dataset+'/'
    os.makedirs(robust_dir, exist_ok=True)
    if option.adjust:
        txt_path = robust_dir+option.model+'_adjust.txt'
        excel_path = robust_dir+option.model+'_adjust.xlsx'
    else:
        txt_path = robust_dir+option.model+'.txt'
        excel_path = robust_dir+option.model+'.xlsx'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    ft = open(txt_path,'w')
    ft.write(option.model+' '+option.dataset+'\n')
    ft.write('Title     Org_acc     Org_gen_acc     FGSM_acc     FGSM_gen_acc     PGD_acc     PGD_gen_acc     Time\n')
    ft.flush()

    print('>'*20 + 'Robustness: '+option.model+' '+option.dataset+'<'*20)

    robust_title_list = []
    org_acc_list = []
    org_gen_acc_list = []
    fgsm_acc_list = []
    fgsm_gen_acc_list = []
    pgd_acc_list = []
    pgd_gen_acc_list = []

    print('>> Full')
    RobustEval = RobustEvaler(option)
    RobustEval.set(False)
    start_time = time.time()
    org_acc, org_gen_acc, fgsm_acc, fgsm_gen_acc, pgd_acc, pgd_gen_acc = RobustEval.run()
    time_interval = time.time()-start_time
    ft.write('Full      %f     %f     %f     %f     %f     %f     %.2fs\n'%(org_acc, org_gen_acc, fgsm_acc, fgsm_gen_acc, pgd_acc, pgd_gen_acc, time_interval))
    ft.flush()
    robust_title_list.append('Full')
    org_acc_list.append(org_acc)
    org_gen_acc_list.append(org_gen_acc)
    fgsm_acc_list.append(fgsm_acc)
    fgsm_gen_acc_list.append(fgsm_gen_acc)
    pgd_acc_list.append(pgd_acc)
    pgd_gen_acc_list.append(pgd_gen_acc)
    
    gol._init()
    quant_type_list = ['INT','POT','FLOAT']
    for quant_type in quant_type_list:
        num_bit_list = numbit_list(quant_type)
        if quant_type != 'INT':
            bias_list = build_bias_list(quant_type)
            gol.set_value(bias_list, is_bias=True)
        for num_bits in num_bit_list:
            e_bit_list = ebit_list(quant_type,num_bits)
            for e_bits in e_bit_list:
                if quant_type == 'FLOAT':
                    title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                else:
                    title = '%s_%d' % (quant_type, num_bits)
                # 设置量化表
                if quant_type != 'INT':
                    plist = build_list(quant_type, num_bits, e_bits)
                    gol.set_value(plist)
                print('>> '+title)
                RobustEval.set(True,quant_type,num_bits,e_bits)
                start_time = time.time()
                org_acc, org_gen_acc, fgsm_acc, fgsm_gen_acc, pgd_acc, pgd_gen_acc = RobustEval.run()
                time_interval = time.time()-start_time
                ft.write(title+'     %f     %f     %f     %f     %f     %f     %.2fs\n'%(org_acc, org_gen_acc, fgsm_acc, fgsm_gen_acc, pgd_acc, pgd_gen_acc, time_interval))
                ft.flush()
                robust_title_list.append(title)
                org_acc_list.append(org_acc)
                org_gen_acc_list.append(org_gen_acc)
                fgsm_acc_list.append(fgsm_acc)
                fgsm_gen_acc_list.append(fgsm_gen_acc)
                pgd_acc_list.append(pgd_acc)
                pgd_gen_acc_list.append(pgd_gen_acc)


    worksheet.cell(row=1,column=1,value='Title')
    worksheet.cell(row=1,column=2,value='Org_acc')
    worksheet.cell(row=1,column=3,value='Org_gen_acc')
    worksheet.cell(row=1,column=4,value='FGSM_acc')
    worksheet.cell(row=1,column=5,value='FGSM_gen_acc')
    worksheet.cell(row=1,column=6,value='PGD_acc')
    worksheet.cell(row=1,column=7,value='PGD_gen_acc')
    for i in range(len(robust_title_list)):
        worksheet.cell(row=i+3,column=1,value=robust_title_list[i])
        worksheet.cell(row=i+3,column=2,value=org_acc_list[i])
        worksheet.cell(row=i+3,column=3,value=org_gen_acc_list[i])
        worksheet.cell(row=i+3,column=4,value=fgsm_acc_list[i])
        worksheet.cell(row=i+3,column=5,value=fgsm_gen_acc_list[i])
        worksheet.cell(row=i+3,column=6,value=pgd_acc_list[i])
        worksheet.cell(row=i+3,column=7,value=pgd_gen_acc_list[i])
    
    workbook.save(excel_path)
    ft.close()


if __name__ == '__main__':
    main()
