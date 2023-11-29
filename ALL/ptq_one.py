from torch.serialization import load
from model import *
from extract_ratio import *
from utils import *
from dataloader import DataLoader

import gol
import openpyxl
import sys
import torch
import os
import os.path as osp


def direct_quantize(model, test_loader,device):
    with torch.no_grad():
        for i, (data, target) in enumerate(test_loader,1):
            data = data.to(device)
            output = model.quantize_forward(data)
            if i % 100 == 0:
                break
    print('direct quantization finish')


def full_inference(model, test_loader, device):
    correct = 0
    with torch.no_grad():
        for data, target in test_loader:
            data,target = data.to(device), target.to(device)
            output = model(data)
            pred = output.argmax(dim=1, keepdim=True)
            # print(pred)
            correct += pred.eq(target.view_as(pred)).sum().item()
    print('\nTest set: Full Model Accuracy: {:.2f}%'.format(100. * correct / len(test_loader.dataset)))
    return 100. * correct / len(test_loader.dataset)


def quantize_inference(model, test_loader, device):
    correct = 0
    with torch.no_grad():
        for data, target in test_loader:
            data,target = data.to(device), target.to(device)
            output = model.quantize_inference(data)
            pred = output.argmax(dim=1, keepdim=True)
            correct += pred.eq(target.view_as(pred)).sum().item()
    print('Test set: Quant Model Accuracy: {:.2f}%'.format(100. * correct / len(test_loader.dataset)))
    return 100. * correct / len(test_loader.dataset)


if __name__ == "__main__":
    batch_size = 128
    model_name = sys.argv[1]
    dataset = sys.argv[2]

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(device)

    dataloader = DataLoader(dataset,batch_size)
    train_loader, val_loader, test_loader = dataloader.getloader()

    load_ptq = False
    store_ptq = True

    append = False
    gol._init()

    ckpt_full_path = 'ckpt_full/'+dataset
    ckpt_quant_path = 'ckpt_quant/'+dataset
    ptq_result_path = 'ptq_result/'+dataset

    if not osp.exists(ckpt_quant_path) and store_ptq:
        os.makedirs(ckpt_quant_path)
    if not osp.exists(ptq_result_path):
        os.makedirs(ptq_result_path)
    
    excel_path = ptq_result_path+'/'+model_name+'.xlsx'
    txt_path = ptq_result_path+'/'+model_name+'.txt'

    if append and os.path.exists(excel_path) and os.path.exists(txt_path):
        print("> Exit: "+model_name+"has been quantized")
        sys.exit()
    else:
        workbook = openpyxl.Workbook()
        ft = open(txt_path,'w')
    
    full_file = ckpt_full_path+'/'+model_name+'.pt'

    ptq_file_dir = ckpt_quant_path+'/'+model_name
    if not osp.exists(ptq_file_dir):
        os.makedirs(ptq_file_dir)

    model = Model(model_name,dataset)
    model.load_state_dict(torch.load(full_file))
    model.to(device)
    model.eval()
    full_acc = full_inference(model, test_loader, device)
    # 传入后可变
    fold_model(model)

    Mac,Param,layer, par_ratio, flop_ratio = extract_ratio(model_name,dataset)
    par_ratio, flop_ratio = fold_ratio(layer, par_ratio, flop_ratio)

    full_names = []
    full_params = []

    for name, param in model.named_parameters():
        if 'conv' in name or 'fc' in name:
            full_names.append(name)
            full_params.append(param.data.cpu())

    quant_type_list = ['INT','POT','FLOAT']
    title_list = []
    js_flops_list = []
    js_flops_wt_log_list = []
    js_flops_wt_cbrt_list = []
    js_param_list = []
    ptq_acc_list = []
    acc_loss_list = []

    for quant_type in quant_type_list:
        num_bit_list = numbit_list(quant_type)
        # 对一个量化类别，只需设置一次bias量化表
        # int由于位宽大，使用量化表开销过大，直接_round即可
        if quant_type != 'INT':
            bias_list = build_bias_list(quant_type)
            gol.set_value(bias_list, is_bias=True)

        for num_bits in num_bit_list:
            e_bit_list = ebit_list(quant_type,num_bits)
            for e_bits in e_bit_list:
                model_ptq = Model(model_name,dataset)
                if quant_type == 'FLOAT':
                    title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                else:
                    title = '%s_%d' % (quant_type, num_bits)
                print('\n'+model_name+': PTQ: '+title)
                title_list.append(title)

                # 设置量化表
                if quant_type != 'INT':
                    plist = build_list(quant_type, num_bits, e_bits)
                    gol.set_value(plist)
                # 判断是否需要载入
                ptq_file = ptq_file_dir +'/' + title + '.pt'
                if load_ptq is True and osp.exists(ptq_file):
                    model_ptq.quantize(quant_type,num_bits,e_bits)
                    model_ptq.load_state_dict(torch.load(ptq_file))
                    model_ptq.to(device)
                    model_ptq.eval()
                    print('Successfully load ptq model: ' + title)
                else:
                    model_ptq.load_state_dict(torch.load(full_file))
                    model_ptq.to(device)
                    model_ptq.quantize(quant_type,num_bits,e_bits)
                    model_ptq.eval()
                    direct_quantize(model_ptq, train_loader, device)
                    if store_ptq:
                        torch.save(model_ptq.state_dict(), ptq_file)

                model_ptq.freeze()
                ptq_acc = quantize_inference(model_ptq, test_loader, device)
                ptq_acc_list.append(ptq_acc)
                acc_loss = (full_acc - ptq_acc) / full_acc
                acc_loss_list.append(acc_loss)

                #将量化后分布反量化到全精度相同的scale
                model_ptq.fakefreeze()
                # 获取计算量/参数量下的js-div
                js_flops = 0.
                js_param = 0.
                for name, param in model_ptq.named_parameters():
                    if 'conv' not in name and 'fc' not in name:
                        continue
                        
                    prefix = name.rsplit('.',1)[0]
                    layer_idx = layer.index(prefix)
                    name_idx = full_names.index(name)

                    layer_idx = layer.index(prefix)
                    ptq_param = param.data.cpu()

                    js = js_div(ptq_param,full_params[name_idx])

                    js = js.item()
                    if js < 0.:
                        js = 0.
                    js_flops = js_flops + js * flop_ratio[layer_idx]
                    js_param = js_param + js * par_ratio[layer_idx]
                js_flops_wt_log = js_flops * torch.log10(torch.tensor(Mac)).item()
                js_flops_wt_cbrt = js_flops * torch.pow(torch.tensor(Mac),1/3).item()
                js_flops_list.append(js_flops)
                js_flops_wt_log_list.append(js_flops_wt_log)
                js_flops_wt_cbrt_list.append(js_flops_wt_cbrt)
                js_param_list.append(js_param)

                print(title + ': js_flops: %f js_flops_wt_log: %f js_flops_wt_cbrt: %f js_param: %f acc_loss: %f' % (js_flops,js_flops_wt_log, js_flops_wt_cbrt, js_param, acc_loss))
    
    # 写入xlsx
    worksheet = workbook.active
    worksheet.cell(row=1,column=1,value='FP32-acc')
    worksheet.cell(row=1,column=2,value=full_acc)
    worksheet.cell(row=1,column=3,value='Mac')
    worksheet.cell(row=1,column=4,value=Mac)
    worksheet.cell(row=1,column=5,value='Param')
    worksheet.cell(row=1,column=6,value=Param)

    worksheet.cell(row=3,column=1,value='title')
    worksheet.cell(row=3,column=2,value='js_flops')
    worksheet.cell(row=3,column=3,value='js_flops_wt_log')
    worksheet.cell(row=3,column=4,value='js_flops_wt_cbrt')
    worksheet.cell(row=3,column=5,value='js_param')
    worksheet.cell(row=3,column=6,value='ptq_acc')
    worksheet.cell(row=3,column=7,value='acc_loss')
    for i in range(len(title_list)):
        worksheet.cell(row=i+4, column=1, value=title_list[i])
        worksheet.cell(row=i+4, column=2, value=js_flops_list[i])
        worksheet.cell(row=i+4, column=3, value=js_flops_wt_log_list[i])
        worksheet.cell(row=i+4, column=4, value=js_flops_wt_cbrt_list[i])
        worksheet.cell(row=i+4, column=5, value=js_param_list[i])
        worksheet.cell(row=i+4, column=6, value=ptq_acc_list[i])
        worksheet.cell(row=i+4, column=7, value=acc_loss_list[i])

    workbook.save(excel_path)
    
    print(model_name,file=ft)
    print('Full_acc: %f'%full_acc,file=ft)
    print('title_list:',file=ft)
    print(title_list,file=ft)
    print('js_flops_list:',file=ft)
    print(js_flops_list, file=ft)
    print('js_flops_wt_log_list:',file=ft)
    print(js_flops_wt_log_list, file=ft)
    print('js_flops_wt_cbrt_list:',file=ft)
    print(js_flops_wt_cbrt_list, file=ft)
    print('js_param_list:',file=ft)
    print(js_param_list, file=ft)
    print('ptq_acc_list:',file=ft)
    print(ptq_acc_list, file=ft)
    print('acc_loss_list:',file=ft)
    print(acc_loss_list, file=ft)
    print("\n",file=ft)

    ft.close()

