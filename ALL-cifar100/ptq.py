from torch.serialization import load
from model import *
from extract_ratio import *
from utils import *

import gol
import openpyxl
import sys
import argparse
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torchvision import datasets, transforms
from torchvision.transforms.functional import InterpolationMode
import torch.utils.bottleneck as bn
import os
import os.path as osp
from torch.utils.tensorboard import SummaryWriter


def direct_quantize(model, test_loader,device):
    for i, (data, target) in enumerate(test_loader, 1):
        data = data.to(device)
        output = model.quantize_forward(data).cpu()
        if i % 100 == 0:
            break
    print('direct quantization finish')


def full_inference(model, test_loader, device):
    correct = 0
    for i, (data, target) in enumerate(test_loader, 1):
        data = data.to(device)
        output = model(data).cpu()
        pred = output.argmax(dim=1, keepdim=True)
        # print(pred)
        correct += pred.eq(target.view_as(pred)).sum().item()
    print('\nTest set: Full Model Accuracy: {:.2f}%'.format(100. * correct / len(test_loader.dataset)))
    return 100. * correct / len(test_loader.dataset)


def quantize_inference(model, test_loader, device):
    correct = 0
    for i, (data, target) in enumerate(test_loader, 1):
        data = data.to(device)
        output = model.quantize_inference(data).cpu()
        pred = output.argmax(dim=1, keepdim=True)
        correct += pred.eq(target.view_as(pred)).sum().item()
    print('Test set: Quant Model Accuracy: {:.2f}%'.format(100. * correct / len(test_loader.dataset)))
    return 100. * correct / len(test_loader.dataset)


if __name__ == "__main__":
    batch_size = 128

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(device)
    train_loader = torch.utils.data.DataLoader(
    datasets.CIFAR100('/lustre/datasets/CIFAR100', train=True, download=False,
                     transform=transforms.Compose([
                        transforms.RandomCrop(32, padding=2),
                        transforms.RandomHorizontalFlip(),
                        transforms.ToTensor(),
                        transforms.Normalize(
                            (0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))
                        ])),
     batch_size=batch_size, shuffle=True, num_workers=4, pin_memory=True
    )

    test_loader = torch.utils.data.DataLoader(
    datasets.CIFAR100('/lustre/datasets/CIFAR100', train=False, download=False, 
                     transform=transforms.Compose([
                            transforms.ToTensor(),
                            transforms.Normalize((0.4914, 0.4822, 0.4465),
                                                (0.2023, 0.1994, 0.2010))
                            ])),
    batch_size=batch_size, shuffle=False, num_workers=4, pin_memory=True
    )

    load_ptq = False
    store_ptq = False

    append = True
    gol._init()

    # ckpt_path = 'ckpt_sgd_5_momentum_95'
    # excel_path = 'ptq_result_sgd_mom95.xlsx'
    # optim_type = 'sgd_5_momentum_95_ntb'
    opt_path = 'adam_lr0001'
    ckpt_path = 'ckpt_'+opt_path
    
    excel_path = 'ptq_result_'+opt_path+'.xlsx'
    if os.path.exists(excel_path) and append:
        workbook = openpyxl.load_workbook(excel_path)
    else:
        workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    # txt_path = 'ptq_result_sgd_mom95.txt'
    txt_path = 'ptq_result_'+opt_path+'.txt'
    if os.path.exists(txt_path) and append:
        ft = open(txt_path,'a')
    else:
        ft = open(txt_path,'w')
    model_name_list = ['AlexNet','AlexNet_BN','VGG_16','VGG_19','Inception_BN',
                       'MobileNetV2','ResNet_18','ResNet_50','ResNet_152']
    # model_name_list = ['AlexNet','AlexNet_BN','VGG_16','VGG_19','Inception_BN',
    #                    'MobileNetV2','ResNet_18','ResNet_50']
    # model_name_list = ['MobileNetV2']
    for model_name in model_name_list:
        if model_name in workbook.sheetnames:
            continue

        model = Model(model_name)
        full_file = ckpt_path+'/cifar100_'+model_name+'.pt'
        model.load_state_dict(torch.load(full_file))
        model.to(device)

        ptq_file_prefix = ckpt_path+'/cifar100_'+model_name+'_ptq_'

        model.eval()
        full_acc = full_inference(model, test_loader, device)
        # 传入后可变
        fold_model(model)
    
        Mac,Param,layer, par_ratio, flop_ratio = extract_ratio(model_name)
        par_ratio, flop_ratio = fold_ratio(layer, par_ratio, flop_ratio)

        full_names = []
        full_params = []

        for name, param in model.named_parameters():
            if 'conv' in name or 'fc' in name:
                full_names.append(name)
                full_params.append(param.data.cpu())
    
        quant_type_list = ['INT','POT','FLOAT']
        # quant_type_list = ['INT']
        # quant_type_list = ['POT']
        # quant_type_list = ['INT','POT']
        title_list = []
        js_flops_list = []
        js_flops_wt_log_list = []
        js_flops_wt_cbrt_list = []
        js_param_list = []
        ptq_acc_list = []
        acc_loss_list = []

        for quant_type in quant_type_list:
            num_bit_list = numbit_list(quant_type)
            # 对一个量化类别，只需设置一次bias量化表
            # int由于位宽大，使用量化表开销过大，直接_round即可
            if quant_type != 'INT':
                bias_list = build_bias_list(quant_type)
                gol.set_value(bias_list, is_bias=True)

            for num_bits in num_bit_list:
                e_bit_list = ebit_list(quant_type,num_bits)
                for e_bits in e_bit_list:
                    model_ptq = Model(model_name)
                    if quant_type == 'FLOAT':
                        title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                    else:
                        title = '%s_%d' % (quant_type, num_bits)
                    print('\n'+model_name+': PTQ: '+title)
                    title_list.append(title)

                    # 设置量化表
                    if quant_type != 'INT':
                        plist = build_list(quant_type, num_bits, e_bits)
                        gol.set_value(plist)
                    # 判断是否需要载入
                    if load_ptq is True and osp.exists(ptq_file_prefix + title + '.pt'):
                        model_ptq.quantize(quant_type,num_bits,e_bits)
                        model_ptq.load_state_dict(torch.load(ptq_file_prefix + title + '.pt'))
                        model_ptq.to(device)
                        print('Successfully load ptq model: ' + title)
                    else:
                        model_ptq.load_state_dict(torch.load(full_file))
                        model_ptq.to(device)
                        model_ptq.quantize(quant_type,num_bits,e_bits)
                        model_ptq.eval()
                        direct_quantize(model_ptq, train_loader, device)
                        if store_ptq:
                            torch.save(model_ptq.state_dict(), ptq_file_prefix + title + '.pt')

                    model_ptq.freeze()
                    ptq_acc = quantize_inference(model_ptq, test_loader, device)
                    ptq_acc_list.append(ptq_acc)
                    acc_loss = (full_acc - ptq_acc) / full_acc
                    acc_loss_list.append(acc_loss)

                    #将量化后分布反量化到全精度相同的scale
                    model_ptq.fakefreeze()
                    # 获取计算量/参数量下的js-div
                    js_flops = 0.
                    js_param = 0.
                    for name, param in model_ptq.named_parameters():
                        if 'conv' not in name and 'fc' not in name:
                            continue
                            
                        prefix = name.rsplit('.',1)[0]
                        layer_idx = layer.index(prefix)
                        name_idx = full_names.index(name)

                        layer_idx = layer.index(prefix)
                        ptq_param = param.data.cpu()

                        js = js_div(ptq_param,full_params[name_idx])

                        js = js.item()
                        if js < 0.:
                            js = 0.
                        js_flops = js_flops + js * flop_ratio[layer_idx]
                        js_param = js_param + js * par_ratio[layer_idx]
                    js_flops_wt_log = js_flops * torch.log10(torch.tensor(Mac)).item()
                    js_flops_wt_cbrt = js_flops * torch.pow(torch.tensor(Mac),1/3).item()
                    js_flops_list.append(js_flops)
                    js_flops_wt_log_list.append(js_flops_wt_log)
                    js_flops_wt_cbrt_list.append(js_flops_wt_cbrt)
                    js_param_list.append(js_param)

                    print(title + ': js_flops: %f js_flops_wt_log: %f js_flops_wt_cbrt: %f js_param: %f acc_loss: %f' % (js_flops,js_flops_wt_log, js_flops_wt_cbrt, js_param, acc_loss))
        
        # 写入xlsx
        worksheet = workbook.create_sheet(model_name)
        worksheet.cell(row=1,column=1,value='FP32-acc')
        worksheet.cell(row=1,column=2,value=full_acc)
        worksheet.cell(row=1,column=3,value='Mac')
        worksheet.cell(row=1,column=4,value=Mac)
        worksheet.cell(row=1,column=5,value='Param')
        worksheet.cell(row=1,column=6,value=Param)

        worksheet.cell(row=3,column=1,value='title')
        worksheet.cell(row=3,column=2,value='js_flops')
        worksheet.cell(row=3,column=3,value='js_flops_wt_log')
        worksheet.cell(row=3,column=4,value='js_flops_wt_cbrt')
        worksheet.cell(row=3,column=5,value='js_param')
        worksheet.cell(row=3,column=6,value='ptq_acc')
        worksheet.cell(row=3,column=7,value='acc_loss')
        for i in range(len(title_list)):
            worksheet.cell(row=i+4, column=1, value=title_list[i])
            worksheet.cell(row=i+4, column=2, value=js_flops_list[i])
            worksheet.cell(row=i+4, column=3, value=js_flops_wt_log_list[i])
            worksheet.cell(row=i+4, column=4, value=js_flops_wt_cbrt_list[i])
            worksheet.cell(row=i+4, column=5, value=js_param_list[i])
            worksheet.cell(row=i+4, column=6, value=ptq_acc_list[i])
            worksheet.cell(row=i+4, column=7, value=acc_loss_list[i])

        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        workbook.save(excel_path)
        
        print(model_name,file=ft)
        print('Full_acc: %f'%full_acc,file=ft)
        print('title_list:',file=ft)
        print(title_list,file=ft)
        print('js_flops_list:',file=ft)
        print(js_flops_list, file=ft)
        print('js_flops_wt_log_list:',file=ft)
        print(js_flops_wt_log_list, file=ft)
        print('js_flops_wt_cbrt_list:',file=ft)
        print(js_flops_wt_cbrt_list, file=ft)
        print('js_param_list:',file=ft)
        print(js_param_list, file=ft)
        print('ptq_acc_list:',file=ft)
        print(ptq_acc_list, file=ft)
        print('acc_loss_list:',file=ft)
        print(acc_loss_list, file=ft)
        print("\n",file=ft)
        ft.flush()

    ft.close()

