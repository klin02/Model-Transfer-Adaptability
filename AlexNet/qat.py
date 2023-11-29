from model import *
from extract_ratio import *
from utils import *

import openpyxl
import gol
import sys
import argparse
import torch
import torch.nn as nn
import torch.optim as optim
from torchvision import datasets, transforms
import os
import os.path as osp
from torch.utils.tensorboard import SummaryWriter


def js_div_norm(a,b):
    a_norm = F.normalize(a.data,p=2,dim=-1)
    b_norm = F.normalize(b.data,p=2,dim=-1)
    return js_div(a_norm,b_norm).cpu().item()


def quantize_aware_training(model, device, train_loader, optimizer, epoch):
    lossLayer = torch.nn.CrossEntropyLoss()
    #统计loss和每个参数的grad
    #初始化
    loss_sum = 0.
    grad_dict = {}
    for name,param in model.named_parameters():
        grad_dict[name] = torch.zeros_like(param) #param.grad和param形状相同
    
    for batch_idx, (data, target) in enumerate(train_loader, 1):
        data, target = data.to(device), target.to(device)
        optimizer.zero_grad()
        output = model.quantize_forward(data)
        # 对一批数据求得的loss是平均值
        loss = lossLayer(output, target)
        loss.backward()
        
        #loss和grads累加
        loss_sum += loss
        for name,param in model.named_parameters():
            if param.grad is not None:
                # print('-------'+name+'-------')
                grad_dict[name] += param.grad
                # print(grad_dict[name])

        # print(grad_dict.items())
        # input()
        optimizer.step()

        if batch_idx % 50 == 0:
            print('Quantize Aware Training Epoch: {} [{}/{}]\tLoss: {:.6f}'.format(
                epoch, batch_idx * len(data), len(train_loader.dataset), loss.item()
            ))
    
    batch_size = len(train_loader.batch_sampler)
    #对不同batch累加值求平均
    for name,grad in grad_dict.items():
        grad_dict[name] = grad / batch_size
    loss_avg = loss_sum / batch_size
    return loss_avg, grad_dict

def full_inference(model, test_loader):
    correct = 0
    for i, (data, target) in enumerate(test_loader, 1):
        data, target = data.to(device), target.to(device)
        output = model(data)
        pred = output.argmax(dim=1, keepdim=True)
        correct += pred.eq(target.view_as(pred)).sum().item()
    print('\nTest set: Full Model Accuracy: {:.2f}%\n'.format(100. * correct / len(test_loader.dataset)))

def train(model, device, train_loader, optimizer, epoch):
    model.train()
    lossLayer = torch.nn.CrossEntropyLoss()
    #统计loss和每个参数的grad
    #初始化
    loss_sum = 0.
    grad_dict = {}
    for name,param in model.named_parameters():
        grad_dict[name] = torch.zeros_like(param) #param.grad和param形状相同
    
    for batch_idx, (data, target) in enumerate(train_loader):
        data, target = data.to(device), target.to(device)
        optimizer.zero_grad()
        output = model(data)
        loss = lossLayer(output, target)
        loss.backward()
        
        #loss和grads累加
        loss_sum += loss
        for name,param in model.named_parameters():
            if param.grad is not None:
                # print('-------'+name+'-------')
                grad_dict[name] += param.grad
                # print(grad_dict[name])        

        optimizer.step()

        if batch_idx % 50 == 0:
            print('Train Epoch: {} [{}/{}]\tLoss: {:.6f}'.format(
                epoch, batch_idx * len(data), len(train_loader.dataset), loss.item()
            ))

    batch_size = len(train_loader.batch_sampler)
    #对不同batch累加值求平均
    for name,grad in grad_dict.items():
        grad_dict[name] = grad / batch_size
    loss_avg = loss_sum / batch_size
    return loss_avg, grad_dict


def quantize_inference(model, test_loader):
    correct = 0
    for i, (data, target) in enumerate(test_loader, 1):
        data, target = data.to(device), target.to(device)
        output = model.quantize_inference(data)
        pred = output.argmax(dim=1, keepdim=True)
        correct += pred.eq(target.view_as(pred)).sum().item()
    print('\nTest set: Quant Model Accuracy: {:.2f}%\n'.format(100. * correct / len(test_loader.dataset)))


if __name__ == "__main__":

    batch_size = 32
    seed = 1
    epochs = 15
    lr = 0.001
    momentum = 0.5

    torch.manual_seed(seed)

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    writer = SummaryWriter(log_dir='./log/qat')
    wb = openpyxl.Workbook()
    ws = wb.active

    layer, par_ratio, flop_ratio = extract_ratio()

    train_loader = torch.utils.data.DataLoader(
        datasets.CIFAR10('../data', train=True, download=True,
                       transform=transforms.Compose([
                            transforms.ToTensor(),
                            transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))
                       ])),
        batch_size=batch_size, shuffle=True, num_workers=1, pin_memory=False
    )

    test_loader = torch.utils.data.DataLoader(
        datasets.CIFAR10('../data', train=False, transform=transforms.Compose([
            transforms.ToTensor(),
            transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))
        ])),
        batch_size=batch_size, shuffle=True, num_workers=1, pin_memory=False
    )


    full_file = 'ckpt/cifar10_AlexNet.pt'
    model = AlexNet()
    # model.load_state_dict(torch.load(full_file))
    model.to(device)
    optimizer = optim.SGD(model.parameters(), lr=lr, momentum=momentum)
    
    load_qat = False
    ckpt_prefix = "ckpt/qat/"

    loss_sum = 0.
    full_grad_sum = {}
    full_grad_avg = {}
    for name,param in model.named_parameters():
        full_grad_sum[name] = torch.zeros_like(param)
        full_grad_avg[name] = torch.zeros_like(param)
    for epoch in range(1, epochs+1):
        # 训练原模型，获取梯度分布
        loss,full_grad = train(model, device, train_loader, optimizer, epoch)
        if epoch == 1:
            loss_start = loss
        # print('loss:%f' % loss_avg)
        writer.add_scalar('Full.loss',loss,epoch)
        # for name,grad in grad_dict.items():
        #     writer.add_histogram('Full.'+name+'_grad',grad,global_step=epoch)

        loss_sum += loss
        loss_avg = loss_sum / epoch        
        loss_delta = loss - loss_start
        for name,grad in full_grad.items():
            full_grad_sum[name] += full_grad[name]
            full_grad_avg[name] = full_grad_sum[name] / epoch
    
        if epoch % 5 == 0:
            ws = wb.create_sheet('epoch_%d'%epoch)
            ws.cell(row=1,column=2,value='loss')
            ws.cell(row=1,column=3,value='loss_sum')
            ws.cell(row=1,column=4,value='loss_avg')
            ws.cell(row=1,column=5,value='loss_delta')
            ws.cell(row=2,column=1,value='FP32')
            ws.cell(row=2,column=2,value=loss.item())
            ws.cell(row=2,column=3,value=loss_sum.item())
            ws.cell(row=2,column=4,value=loss_avg.item())
            ws.cell(row=2,column=5,value=loss_delta.item())
            
            ws.cell(row=4,column=1,value='title')
            ws.cell(row=4,column=2,value='loss')
            ws.cell(row=4,column=3,value='loss_sum')
            ws.cell(row=4,column=4,value='loss_avg')
            ws.cell(row=4,column=5,value='loss_delta')
            ws.cell(row=4,column=6,value='js_grad')
            ws.cell(row=4,column=7,value='js_grad_sum')
            ws.cell(row=4,column=8,value='js_grad_avg')

            

    quant_type_list = ['INT','POT','FLOAT']
    gol._init()
    
    currow=4 #数据从哪行开始写
    for quant_type in quant_type_list:
        num_bit_list = numbit_list(quant_type)

        # 对一个量化类别，只需设置一次bias量化表
        # int由于位宽大，使用量化表开销过大，直接_round即可
        if quant_type != 'INT':
            bias_list = build_bias_list(quant_type)
            gol.set_value(bias_list, is_bias=True)

        for num_bits in num_bit_list:
            e_bit_list = ebit_list(quant_type,num_bits)
            for e_bits in e_bit_list:
                
                if quant_type == 'FLOAT':
                    title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                else:
                    title = '%s_%d' % (quant_type, num_bits)

                if load_qat is True and osp.exists(ckpt_prefix+'epoch_20/'+title+'.pt'):
                    continue

                currow += 1
                print('\nQAT: '+title)
                
                model_ptq = AlexNet()
                 # 设置量化表
                if quant_type != 'INT':
                    plist = build_list(quant_type, num_bits, e_bits)
                    gol.set_value(plist)
                # model_ptq.load_state_dict(torch.load(full_file))
                model_ptq.to(device)
                model_ptq.quantize(quant_type,num_bits,e_bits)
                model_ptq.train()

                loss_sum = 0.
                qat_grad_sum = {}
                qat_grad_avg = {}
                for name,param in model.named_parameters():
                    qat_grad_sum[name] = torch.zeros_like(param)
                    qat_grad_avg[name] = torch.zeros_like(param)
                for epoch in range(1, epochs+1):
                    loss,qat_grad = quantize_aware_training(model_ptq, device, train_loader, optimizer, epoch)
                    # print('loss:%f' % loss_avg)
                    if epoch == 1:
                        loss_start = loss
                    writer.add_scalar(title+'.loss',loss,epoch)
                    # for name,grad in qat_grad.items():
                    #     writer.add_histogram(title+'.'+name+'_grad',grad,global_step=epoch)

                    loss_sum += loss
                    loss_avg = loss_sum / epoch 
                    loss_delta = loss-loss_start       
                    for name,param in model.named_parameters():
                        qat_grad_sum[name] += qat_grad[name]
                        qat_grad_avg[name] = qat_grad_sum[name] / epoch

                    if epoch % 5 == 0:
                        ws = wb['epoch_%d'%epoch]
                        js_grad = 0.
                        js_grad_sum = 0.
                        js_grad_avg = 0.
                        for name,_ in model_ptq.named_parameters():
                            prefix = name.split('.')[0]
                            layer_idx = layer.index(prefix)
                            js_grad += flop_ratio[layer_idx] * js_div_norm(qat_grad[name],full_grad[name])
                            js_grad_sum += flop_ratio[layer_idx] * js_div_norm(qat_grad_sum[name],full_grad_sum[name])
                            js_grad_avg += flop_ratio[layer_idx] * js_div_norm(qat_grad_avg[name],full_grad_avg[name])
                        
                        ws.cell(row=currow,column=1,value=title)
                        ws.cell(row=currow,column=2,value=loss.item())
                        ws.cell(row=currow,column=3,value=loss_sum.item())
                        ws.cell(row=currow,column=4,value=loss_avg.item())
                        ws.cell(row=currow,column=5,value=loss_delta.item())
                        ws.cell(row=currow,column=6,value=js_grad)
                        ws.cell(row=currow,column=7,value=js_grad_sum)
                        ws.cell(row=currow,column=8,value=js_grad_avg)

    wb.remove(wb['Sheet'])  # 根据名称删除工作表
    wb.save('qat_result.xlsx')
    writer.close()
