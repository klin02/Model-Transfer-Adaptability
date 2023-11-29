from model import *
from extract_ratio import *
from utils import *

import openpyxl
import os
import os.path as osp
import torch.nn.functional as F
from torch.utils.tensorboard import SummaryWriter

def js_div_norm(a,b):
    a_norm = F.normalize(a.data,p=2,dim=-1)
    b_norm = F.normalize(b.data,p=2,dim=-1)
    return js_div(a_norm,b_norm)


if __name__ == "__main__":
    wb = openpyxl.Workbook()
    ws = wb.active

    writer = SummaryWriter(log_dir='./log')

    layer, par_ratio, flop_ratio = extract_ratio()

    dir_prefix = 'ckpt/qat/epoch_'
    quant_type_list = ['INT','POT','FLOAT']
    
    for epoch in [5,10,15,20]:
        ws_epoch = wb.create_sheet('epoch_%d'%epoch)
        full_state = torch.load(dir_prefix+'%d/'%epoch + 'full.pt')
        ws_epoch.cell(row=1,column=2,value='loss')
        ws_epoch.cell(row=1,column=3,value='loss_sum')
        ws_epoch.cell(row=1,column=4,value='loss_avg')
        ws_epoch.cell(row=2,column=1,value='FP32')
        ws_epoch.cell(row=2,column=2,value=full_state['loss'].cpu().item())
        ws_epoch.cell(row=2,column=3,value=full_state['loss_sum'].cpu().item())
        ws_epoch.cell(row=2,column=4,value=full_state['loss_avg'].cpu().item())
        # full_grad = full_state['grad_dict']
        # full_grad_sum = full_state['grad_dict_sum']
        full_grad_avg = full_state['grad_dict_avg']
        for name,tmpgrad in full_grad_avg.items():
            writer.add_histogram('FULL: '+name,tmpgrad,global_step=epoch)
        ws_epoch.cell(row=4,column=1,value='title')
        ws_epoch.cell(row=4,column=2,value='loss')
        ws_epoch.cell(row=4,column=3,value='loss_sum')
        ws_epoch.cell(row=4,column=4,value='loss_avg')
        ws_epoch.cell(row=4,column=5,value='js_grad_avg_norm')
        ws_epoch.cell(row=4,column=6,value='conv1.weight')
        ws_epoch.cell(row=4,column=7,value='conv1.bias')
        ws_epoch.cell(row=4,column=8,value='conv2.weight')
        ws_epoch.cell(row=4,column=9,value='conv2.bias')
        ws_epoch.cell(row=4,column=10,value='conv3.weight')
        ws_epoch.cell(row=4,column=11,value='conv3.bias')
        ws_epoch.cell(row=4,column=12,value='conv4.weight')
        ws_epoch.cell(row=4,column=13,value='conv4.bias')
        ws_epoch.cell(row=4,column=14,value='conv5.weight')
        ws_epoch.cell(row=4,column=15,value='conv5.bias')
        ws_epoch.cell(row=4,column=16,value='fc1.weight')
        ws_epoch.cell(row=4,column=17,value='fc1.bias')
        ws_epoch.cell(row=4,column=18,value='fc2.weight')
        ws_epoch.cell(row=4,column=19,value='fc2.bias')
        ws_epoch.cell(row=4,column=20,value='fc3.weight')
        ws_epoch.cell(row=4,column=21,value='fc3.bias')
        currow=4
        for quant_type in quant_type_list:
            num_bit_list = numbit_list(quant_type)
            for num_bits in num_bit_list:
                e_bit_list = ebit_list(quant_type,num_bits)
                for e_bits in e_bit_list:
                    if quant_type == 'FLOAT':
                        title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                    else:
                        title = '%s_%d' % (quant_type, num_bits)
                    print('\nAnalyse: '+title)
                    currow += 1
                    qat_state=torch.load(dir_prefix+'%d/'%epoch+title+'.pt')
                    js_grad_avg_norm=0.
                    grad_avg = qat_state['grad_dict_avg']
                    for name,tmpgrad in grad_avg.items():
                        writer.add_histogram(title+': '+name,tmpgrad,global_step=epoch)
                    
                    colidx=5
                    for name,_ in full_grad_avg.items():
                        prefix = name.split('.')[0]
                        colidx += 1
                        layer_idx = layer.index(prefix)
                        js_norm = js_div_norm(full_grad_avg[name],grad_avg[name])
                        ws_epoch.cell(row=currow,column=colidx,value=js_norm.cpu().item())
                        js_grad_avg_norm += flop_ratio[layer_idx] * js_norm

                    ws_epoch.cell(row=currow,column=1,value=title)
                    ws_epoch.cell(row=currow,column=2,value=qat_state['loss'].cpu().item())
                    ws_epoch.cell(row=currow,column=3,value=qat_state['loss_sum'].cpu().item())
                    ws_epoch.cell(row=currow,column=4,value=qat_state['loss_avg'].cpu().item())
                    ws_epoch.cell(row=currow,column=5,value=js_grad_avg_norm.cpu().item())
        wb.save('loss_grad.xlsx')
    writer.close()